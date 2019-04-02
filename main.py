from selenium import webdriver
import time
from threading import Thread, Timer
import pandas
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup
import smtplib

global LIMIT_DEEP
global SCRAPER_THREAD_MAX_NUM
global URLS_FILENAME
global KEYWORD_FILENAME
global MAX_VISITED_PAGES
global MAIL_SENDER
global MAIL_SENDER_PASSWORD
global MAIL_RECEIVER
global VPS_ID
global MODE

# SCRAPER CONFIG
LIMIT_DEEP = 2
SCRAPER_THREAD_MAX_NUM = 5
MAX_VISITED_PAGES = 150
MODE = "PRO"
# ERR and PRO
LIMITED_TYPE = [".jpg", ".png", ".bmp", ".mp4", ".avi", ".mp3", ".rar", ".zip", ".doc", ".pdf"]
print("Input Urls FileName")
URLS_FILENAME = input()
# FILE PATH FOR BASE DATA
BASE_INPUT_FLODER = "./input/"

# URLS_FILENAME = "50urls.xlsx"
CATEGORY_KEY_FILENAME   = BASE_INPUT_FLODER + "category_keywords.xlsx"
INDUSTRY_KEY_FILENAME   = BASE_INPUT_FLODER + "industry_keywords.xlsx"
MARKET_KEY_FILENAME     = BASE_INPUT_FLODER + "market_keywords.xlsx"
PINTEREST_KEY_FILENAME  = BASE_INPUT_FLODER + "pinterest_keywords.xlsx"

# FILE PATH FOR RESULT
BASE_OUTPUT_FLODER = "./output/"

# split small files
global CATEGORY_RESULT_FILENAME
global INDUSTRY_RESULT_FILENAME
global MARKET_RESULT_FILENAME
global PINTEREST_RESULT_FILENAME
global ERROR_RESULT_FILENAME

CATEGORY_RESULT_FILENAME   = BASE_OUTPUT_FLODER + URLS_FILENAME + "_category_result_"
INDUSTRY_RESULT_FILENAME   = BASE_OUTPUT_FLODER + URLS_FILENAME + "_industry_result_"

MARKET_RESULT_FILENAME     = BASE_OUTPUT_FLODER + URLS_FILENAME + "_market_result.xlsx"
PINTEREST_RESULT_FILENAME  = BASE_OUTPUT_FLODER + URLS_FILENAME + "_pinterest_result.xlsx"
ERROR_RESULT_FILENAME      = BASE_OUTPUT_FLODER + URLS_FILENAME + "_error_urls.xlsx"

URLS_FILENAME = URLS_FILENAME + ".xlsx"
# EMAIL SETTING
MAIL_SENDER = "xxx@gmail.com"
MAIL_RECEIVER = ["Xxx@gmail.com"]
MAIL_SENDER_PASSWORD = "yyy"
VPS_ID = "TEST_ID"

class KeywordDeepScraper(Thread):
    
    def __init__(self, url):
        Thread.__init__(self)
        if MODE == "PRO":
            chromeOptions = webdriver.ChromeOptions()
            prefs = {'profile.managed_default_content_settings.images':2}
            chromeOptions.add_experimental_option("prefs", prefs)
            self.browser = webdriver.Chrome(chrome_options=chromeOptions)
        else: 
            self.browser = webdriver.Chrome()
        self.error = False
        self.url = url
        self.baseurl = url
        self.category_result    = {"url":[], "keyword":[], "count":[]}
        self.industry_result    = {"url":[], "keyword":[], "count":[]}
        self.market_result      = {"url":[], "keyword":[], "count":[]}
        self.pinterest_result   = {"url":[], "keyword":[], "count":[]}
        self.visitedlist        = []

    def crawlPages(self, url, deep):

        if url in self.visitedlist:
            return True
        else:
            self.visitedlist.append(url)

        if deep > LIMIT_DEEP :
            return True
        visitedlen = len(self.visitedlist)
        if visitedlen > MAX_VISITED_PAGES:
            return True

        print("==============>>> Deep:"+str(deep)+" Visited Order:"+str(visitedlen)+"] <<<=============="+url)
        self.browser.get(url)
        # if error or zerodata urls
        if MODE == "PRO":
            time.sleep(2)
        else:
            time.sleep(30)

        soup = BeautifulSoup(self.browser.page_source, "html.parser")
        page = soup.find("body").getText()

        if page.find("This site can’t be reached")>-1:
            print("This site can’t be reached")
            return False
            # raise Exception('This site can’t be reached')

        cur_url = self.browser.current_url
        if deep == 0:
            self.baseurl = cur_url
        elif cur_url.find(self.baseurl) is not 0:
            return True
      
        self.searchKeyword(page, cur_url)

        if deep == LIMIT_DEEP: 
            return True
        else: 
            alist = self.browser.find_elements_by_xpath("//a")
            # print(alist)
            links = []
            if alist:
                for atag in alist:
                    link = atag.get_attribute("href")
                    if link:
                        # limit download href
                        if link[-4:] in LIMITED_TYPE:
                            continue
            
                        if link.find(self.baseurl) == 0:
                            if link[-1] == '/':
                                link = link[:-1]                                
                            if link not in self.visitedlist and link not in links:
                                links.append(link)
                        elif link.find(".") == -1 and link.find("javascript") == -1 and link.find("#") == -1 and link[0] == "/":
                            link = self.baseurl + link
                            if link not in self.visitedlist and link not in links:
                                links.append(link)
                        else:
                            continue
            for link in links:
                # print(link)
                self.crawlPages(link, deep + 1)
            return True

    def searchKeyword(self, page, url):
        try:
            page = page.lower()
            # print(page)
            for keyword in CATEGORY_KEYWORDS:
                count = page.count(keyword)
                if count:
                    self.category_result['url'].append(url)
                    self.category_result['keyword'].append(keyword)
                    self.category_result['count'].append(count)

            for keyword in INDUSTRY_KEYWORDS:
                count = page.count(keyword)
                if count:
                    self.industry_result['url'].append(url)
                    self.industry_result['keyword'].append(keyword)
                    self.industry_result['count'].append(count)
            for keyword in MARKET_KEYWORDS:
                count = page.count(keyword)
                if count:
                    self.market_result['url'].append(url)
                    self.market_result['keyword'].append(keyword)
                    self.market_result['count'].append(count)

            for keyword in PINTEREST_KEYWORDS:
                count = page.count(keyword)
                if count:
                    self.pinterest_result['url'].append(url)
                    self.pinterest_result['keyword'].append(keyword)
                    self.pinterest_result['count'].append(count)
        except:
            print("Search Error")
            raise Exception('Search Error')

    def test(self):
        for i in range(100):
            time.sleep(.1)
            print(self.url)
    
    def addZeroRow(self):
        self.category_result['url'].append(self.url)
        self.category_result['keyword'].append(0)
        self.category_result['count'].append(0)
        self.industry_result['url'].append(self.url)
        self.industry_result['keyword'].append(0)
        self.industry_result['count'].append(0)
        self.market_result['url'].append(self.url)
        self.market_result['keyword'].append(0)
        self.market_result['count'].append(0)
        self.pinterest_result['url'].append(self.url)
        self.pinterest_result['keyword'].append(0)
        self.pinterest_result['count'].append(0)
            
    def run(self):
        # self.timer.start()
        start_time = time.time()
        try:
            self.crawlPages(self.url, 0)
        except:
            print("ERROR from " + self.url)
        
        if len(self.category_result['url']) > 0 or len(self.industry_result['url']) > 0:
            self.addZeroRow()
        else:
            self.error = True

        duration = "%s seconds" % (time.time()-start_time)
        visitedlen = len(self.visitedlist)
        print(">>>>>>>FNISHED>>>>>>\t"+self.url+">>>>"+str(visitedlen)+"PAGES>>>>\t"+duration)
        self.browser.quit()

    def timeout(self):
        print("timeout")
        return False

def readKeywords():

    url = pandas.read_excel(URLS_FILENAME, dtype={'URLS':str})
    c_k = pandas.read_excel(CATEGORY_KEY_FILENAME, dtype={'CATEGORY_KEYWORDS':str})
    i_k = pandas.read_excel(INDUSTRY_KEY_FILENAME, dtype={'INDUSTRY_KEYWORDS':str})
    m_k = pandas.read_excel(MARKET_KEY_FILENAME, dtype={'MARKET_KEYWORDS':str})
    p_k = pandas.read_excel(PINTEREST_KEY_FILENAME, dtype={'PINTEREST_KEYWORDS':str})
    
    global URLS
    global CATEGORY_KEYWORDS
    global INDUSTRY_KEYWORDS
    global MARKET_KEYWORDS
    global PINTEREST_KEYWORDS

    URLS = url['URLS']
    CATEGORY_KEYWORDS   = c_k['CATEGORY_KEYWORDS']
    INDUSTRY_KEYWORDS   = i_k['INDUSTRY_KEYWORDS']
    MARKET_KEYWORDS     = m_k['MARKET_KEYWORDS']
    PINTEREST_KEYWORDS  = p_k['PINTEREST_KEYWORDS']


def writeResult(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pandas.ExcelWriter(filename, engine='openpyxl')
    
    try:
        writer.book = load_workbook(filename)
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
       
    except:
        # print("writeError")
        pass
    if startrow is None:
        startrow = 0

    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs, header=False)
    writer.save()

def Notify():
    message = " SCRAPING FNISHED.PLESE CHECK" + "<h1>"+ VPS_ID +"</h1>"
    try:
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.login(MAIL_SENDER, MAIL_SENDER_PASSWORD)
        server.sendmail(
            MAIL_SENDER, 
            MAIL_RECEIVER, 
            message
        )
        server.quit()
    except Exception as e:
        print(e)

def main():
        
    readKeywords()
    start_time = time.time()
    urls_num = len(URLS)
    pack_num = int(urls_num / SCRAPER_THREAD_MAX_NUM) + 1

    for i in range(pack_num):
        print("*******************************************")
        print("*******************>" + str(i) + "<*********************")
        print("*******************************************")
        threads     = []
        start_index = i * SCRAPER_THREAD_MAX_NUM
        end_index   = (i + 1) * SCRAPER_THREAD_MAX_NUM
        fileindex = int(i/2)
              
        P_CATEGORY_RESULT_FILENAME = CATEGORY_RESULT_FILENAME + str(fileindex) + ".xlsx"
        P_INDUSTRY_RESULT_FILENAME = INDUSTRY_RESULT_FILENAME + str(fileindex) + ".xlsx"
        if end_index > urls_num:
            end_index = urls_num
        for url in URLS[start_index: end_index]:
            if MODE == "PRO":
                url =  "http://" + url
            thread = KeywordDeepScraper(url)
            threads.append(thread)
            thread.start()

        for thread in threads:
            thread.join()
            
        for thread in threads:
            if thread.error:
                df = pandas.DataFrame({'Urls': [thread.url]})
                writeResult(ERROR_RESULT_FILENAME, df)
            else:
                c_df = pandas.DataFrame(thread.category_result)
                i_df = pandas.DataFrame(thread.industry_result)
                m_df = pandas.DataFrame(thread.market_result)
                p_df = pandas.DataFrame(thread.pinterest_result)           
                writeResult(P_CATEGORY_RESULT_FILENAME, c_df)
                writeResult(P_INDUSTRY_RESULT_FILENAME, i_df)
                writeResult(MARKET_RESULT_FILENAME, m_df)
                writeResult(PINTEREST_RESULT_FILENAME, p_df)
            del thread
    print("--- %s seconds ---" % (time.time()-start_time))

def test():
    readKeywords()
    start_time = time.time()
    urls = ['http://www.digitalprosoft.com/', 'http://directlawstrategies.com', 'https://doctual.com/']
    threads = []
    for url in urls:        
        thread = KeywordDeepScraper(url)
        threads.append(thread)
        thread.start()
    for thread in threads:
        thread.join()

    for thread in threads:
        if thread.isAlive():
            print("LIVE")
        else:
            print("DIE")
            print(thread.category_result)
            print(thread.industry_result)
            # print(thread.market_result)
            # print(thread.pinterest_result)
    
    print("--- %s seconds ---" % (time.time()-start_time))


if __name__ == "__main__":
    main()
    # test()
