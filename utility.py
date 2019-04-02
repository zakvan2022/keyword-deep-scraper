import pandas

BASE_INPUT_FLODER = "./input/"
CATEGORY_KEY_FILENAME   = BASE_INPUT_FLODER + "category_keywords.xlsx"
INDUSTRY_KEY_FILENAME   = BASE_INPUT_FLODER + "industry_keywords.xlsx"
MARKET_KEY_FILENAME     = BASE_INPUT_FLODER + "market_keywords.xlsx"
PINTEREST_KEY_FILENAME  = BASE_INPUT_FLODER + "pinterest_keywords.xlsx"

CATEGORY_RESULT_FILENAME   = BASE_INPUT_FLODER + "category_filtered_keywords.xlsx"
INDUSTRY_RESULT_FILENAME   = BASE_INPUT_FLODER + "industry_filtered_keywords.xlsx"
MARKET_RESULT_FILENAME     = BASE_INPUT_FLODER + "market_filtered_keywords.xlsx"
PINTEREST_RESULT_FILENAME  = BASE_INPUT_FLODER + "pinterest_filtered_keywords.xlsx"

URL_FOLDER = "./urls/"
URLS_FILENAME = "urls.xlsx"

def readKeywords():
    c_k = pandas.read_excel(CATEGORY_KEY_FILENAME, dtype={'CATEGORY_KEYWORDS':str})
    i_k = pandas.read_excel(INDUSTRY_KEY_FILENAME, dtype={'INDUSTRY_KEYWORDS':str})
    m_k = pandas.read_excel(MARKET_KEY_FILENAME, dtype={'MARKET_KEYWORDS':str})
    p_k = pandas.read_excel(PINTEREST_KEY_FILENAME, dtype={'PINTEREST_KEYWORDS':str})

    global CATEGORY_KEYWORDS
    global INDUSTRY_KEYWORDS
    global MARKET_KEYWORDS
    global PINTEREST_KEYWORDS

    CATEGORY_KEYWORDS   = c_k['CATEGORY_KEYWORDS']
    INDUSTRY_KEYWORDS   = i_k['INDUSTRY_KEYWORDS']
    MARKET_KEYWORDS     = m_k['MARKET_KEYWORDS']
    PINTEREST_KEYWORDS  = p_k['PINTEREST_KEYWORDS']

def readUrls():
     url = pandas.read_excel(URLS_FILENAME, dtype={'URLS':str})
     global URLS
     URLS = url['URLS']

def filter(oldkeywords, outfilename, tagname):
    filteredkeywords = []
    for keyword in oldkeywords:
        if keyword in filteredkeywords:
            continue
        else:
            filteredkeywords.append(keyword)
    out = pandas.DataFrame({tagname: filteredkeywords})
    writeResult(outfilename, out)

def writeResult(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pandas.ExcelWriter(filename, engine='openpyxl')
    
    try:
        writer.book = load_workbook(filename)
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
       
    except:
        print("WriteError")
        pass
    if startrow is None:
        startrow = 0

    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs, header=True)
    writer.save()

def keywordFilter():
    readKeywords()
    filter(CATEGORY_KEYWORDS,   CATEGORY_RESULT_FILENAME, 'CATEGORY_KEYWORDS')
    filter(INDUSTRY_KEYWORDS,   INDUSTRY_RESULT_FILENAME, 'INDUSTRY_KEYWORDS')
    filter(MARKET_KEYWORDS,     MARKET_RESULT_FILENAME,   'MARKET_KEYWORDS')
    filter(PINTEREST_KEYWORDS,  PINTEREST_RESULT_FILENAME,'PINTEREST_KEYWORDS')

def urlSpliter():
    URL_NUM_PER_FILE = 500
    readUrls()
    urls_num = len(URLS)
    pack_num = int(urls_num / URL_NUM_PER_FILE) + 1
    for i in range(pack_num):
        start_index = i * URL_NUM_PER_FILE
        end_index   = (i + 1) * URL_NUM_PER_FILE
        if end_index > urls_num:
            end_index = urls_num
        urls = URLS[start_index: end_index]
        filename = URL_FOLDER + str(start_index) + "-" + str(end_index) + ".xlsx"
        out = pandas.DataFrame({"URLS": urls})
        writeResult(filename, out)

def filterZeroData(base_folder):
    import glob
    import openpyxl
    from openpyxl import Workbook
    filenames = base_folder + "/*.xlsx"
    filenames = glob.glob(filenames)
    
    save_filename = base_folder + "/zerodata_urls.xlsx"
    savebook = Workbook()
    savesheet = savebook.active
    try:
        for filename in filenames:
            if "category_result" in filename:
                book = openpyxl.load_workbook(filename, data_only=True)
                sheet = book.active
                rows = sheet.rows
                zero_point = -1
                i = 0
                for row in rows:
                    if row[2].value is 0 and row[3].value is 0:
                        # print(row[1].value)
                        if  (zero_point + 1) == i:
                            savesheet.append((row[1].value, filename))
                        zero_point = i

                    i += 1
                print(filename+" is processed")
        savebook.save(save_filename)
        print("Completed successfully")
    except:
        print("Error!!!")

def composeFiles(base_folder):
    import glob
    import openpyxl
    from openpyxl import Workbook

    filenames = base_folder + "/*.xlsx"
    e_filename = base_folder + "/error_total_urls.xlsx"
    c_filename = base_folder + "/a_category_result.xlsx"
    i_filename = base_folder + "/a_industry_result.xlsx"
    i2_filename = base_folder + "/a_industry_result2.xlsx"
    filenames = glob.glob(filenames)

    e_savebook = Workbook()
    e_savesheet = e_savebook.active
    c_savebook = Workbook()
    c_savesheet = c_savebook.active
    i_savebook = Workbook()
    i_savesheet = i_savebook.active
    i2_savebook = Workbook()
    i2_savesheet = i2_savebook.active
    # print(filenames)
    # return
    flag = False
    maxrow = 1000000
    rownum = 0
    try:
        c_i = 0
        i_i = 0
        e_i = 0
        for filename in filenames:
            if "error_urls" in filename:

                book = openpyxl.load_workbook(filename, data_only=True)
                sheet = book.active
                rows = sheet.rows
                for row in rows:
                    if row[0].value == None or row[1].value == None:
                        break
                    e_savesheet.append((row[0].value, row[1].value))
                e_i += 1
                print(filename+" is added")
            
            if "category_result" in filename:

                book = openpyxl.load_workbook(filename, data_only=True)
                sheet = book.active
                rows = sheet.rows
                for row in rows:
                    if row[0].value == None or row[1].value == None or row[2].value == None or row[3].value == None:
                        break
                    c_savesheet.append((row[0].value, row[1].value, row[2].value, row[3].value))
                c_i += 1
                print(filename+" is added")
            
            if "industry_result" in filename:
                book = openpyxl.load_workbook(filename, data_only=True)
                sheet = book.active
                rows = sheet.rows
                for row in rows:
                    if row[0].value == None or row[1].value == None or row[2].value == None or row[3].value == None:
                        break
                    rownum += 1
                    if rownum < maxrow:
                        i_savesheet.append((row[0].value, row[1].value, row[2].value, row[3].value))
                    else:
                        flag = True
                        i2_savesheet.append((row[0].value, row[1].value, row[2].value, row[3].value))
                i_i += 1
                print(filename+" is added")
        e_savebook.save(e_filename)
        c_savebook.save(c_filename)
        i_savebook.save(i_filename)
        if flag:
            i2_savebook.save(i2_filename)
        print("Completed successfully")
    except:
        print("Error!!!")

def xlsxtocsv(base_path):
    import csv
    import glob
    import openpyxl
    from os import listdir
    from os.path import isfile, join
    from openpyxl import Workbook


    catfile = open(join(base_path, "category_result.csv"), "w", newline="", encoding="utf-8")
    category_file = csv.writer(catfile, delimiter=',', quoting=csv.QUOTE_MINIMAL)
    indfile = open(join(base_path, "industry_result.csv"), "w", newline="", encoding="utf-8")
    industry_file = csv.writer(indfile, delimiter=',', quoting=csv.QUOTE_MINIMAL)
    marfile = open(join(base_path, "market_result.csv"), "w", newline="", encoding="utf-8")
    market_file = csv.writer(marfile, delimiter=',', quoting=csv.QUOTE_MINIMAL)
    pinfile = open(join(base_path, "pinterest_result.csv"), "w", newline="", encoding="utf-8")
    pinterest_file = csv.writer(pinfile, delimiter=',', quoting=csv.QUOTE_MINIMAL)
    errfile = open(join(base_path, "error_urls.csv"), "w", newline="", encoding="utf-8")
    error_file = csv.writer(errfile, delimiter=',', quoting=csv.QUOTE_MINIMAL)

    onlydirs = [d for d in listdir(base_path) if not isfile(join(base_path, d))]
    for directory in onlydirs:
        filenames = glob.glob(join(base_path, directory, "*.xlsx"))
        for filename in filenames:
            book = openpyxl.load_workbook(filename, data_only=True)
            sheet = book.active
            rows = sheet.rows
            for row in rows:
                if row[0].value == None:
                    break
                if "category_result" in filename:
                    category_file.writerow([row[0].value, row[1].value, row[2].value, row[3].value])
                elif "industry_result" in filename:
                    industry_file.writerow([row[0].value, row[1].value, row[2].value, row[3].value])
                elif "market_result" in filename:
                    market_file.writerow([row[0].value, row[1].value, row[2].value, row[3].value])
                elif "pinterest_result" in filename:
                    pinterest_file.writerow([row[0].value, row[1].value, row[2].value, row[3].value])
                elif "error_urls" in filename:
                    error_file.writerow([row[0].value, row[1].value])
            print(filename+" is added")
        print(directory+' is completed')
    print("DONE!")



def main():
    print("please input folder name")
    folder = input()
    xlsxtocsv(folder)

if __name__ == "__main__":
    main()